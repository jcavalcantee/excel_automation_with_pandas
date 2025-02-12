import os
import pandas as pd
from openpyxl import load_workbook
from utils import retornaMesAtual, retornaNumeroMes
from calcTempo import medir_tempo, formatar_tempo

regionaisCodigo = [1,2,3,4]
regionaisNome = ['REGIONAL 1', 'REGIONAL 2', 'REGIONAL 3', 'REGIONAL 4']

regionaisCodigo1 = [1,2,3,4]
regionaisNome1 = ['REGIONAL 1', 'REGIONAL 2', 'REGIONAL 3', 'REGIONAL 4']

regionaisCodigo2 = [1,2,3,4]
regionaisNome2 = ['REGIONAL 1', 'REGIONAL 2', 'REGIONAL 3', 'REGIONAL 4']

def arquivoBCMaisRecente():
    diretorio = r'DIRETORIO AQUI'
    arquivos_excel = [f for f in os.listdir(diretorio) if f.endswith('.xlsx') or f.endswith('.xls')]
    arquivo_mais_recente = max(arquivos_excel, key=lambda f: os.path.getmtime(os.path.join(diretorio, f)))
    caminho_arquivo = os.path.join(diretorio, arquivo_mais_recente)
    return pd.read_excel(caminho_arquivo)
        
def filtroBaseConsolidada(filtro):
    df = arquivoBCMaisRecente()
    df_filtrado = df[df['FILTRO AQUI'] == filtro]
    df_final = df_filtrado.copy()
    return df_final

def inserirDadosExcel(filtro, caminho_pasta_dinamica):
    df_filtrado = filtroBaseConsolidada(filtro)
    if df_filtrado.empty:
        print(f"Nenhum dado para o FILTRO {filtro}. Abandonando a operação.")
        return

    print(f"Inserindo os dados filtrados (Org. {filtro})...")
    df_modelo_com_dados = df_filtrado  

    # Carregar o arquivo modelo com openpyxl para preservar a formatação
    caminho_arquivo = r'data/arquivo_modelo.xlsx'
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook['NOME WORKBOOK AQUI']

    start_row = 4

    # Inserir os dados filtrados a partir da linha 4
    for i, row in enumerate(df_modelo_com_dados.values, start=start_row):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)

    # Salvar o arquivo atualizado, preservando o modelo para futuros arquivos
    mes_atual = retornaMesAtual()
    nome_arquivo_formatado = f"ARQUIVO - {filtro} ({mes_atual}'25)"
    
    novo_arquivo_formatado = os.path.join(caminho_pasta_dinamica, f'{nome_arquivo_formatado}.xlsx')
    workbook.save(novo_arquivo_formatado)

    return(f'Arquivo atualizado com os dados do FILTRO {filtro} salvo em: {novo_arquivo_formatado}')

@medir_tempo
def criarPastasRegionais(filtro):
    diretorio_base = r'DIRETORIO AQUI\DINAMICO {}\2025'
    diretorio_especifico_regional = diretorio_base.format(filtro)
    
    numero_mes = retornaNumeroMes()
    mes_ano = retornaMesAtual()
    mes_atual = f'{numero_mes}. {mes_ano}'

    caminho_mes = os.path.join(diretorio_especifico_regional, mes_atual)

    if not os.path.exists(caminho_mes):
        os.makedirs(caminho_mes)
        print(f'Pasta do mês criada: {caminho_mes}')
    else:
        print(f'A pasta do mês já existe: {caminho_mes}')

    if filtro == 'FILTRO A':
        for codigo, nome in zip(regionaisCodigo, regionaisNome):
            nome_pasta = f'{codigo} - {nome}'
            caminho_pasta = os.path.join(caminho_mes, nome_pasta)
            if not os.path.exists(caminho_pasta):
                os.makedirs(caminho_pasta)
                print(f'Pasta criada: {caminho_pasta}')
            else:
                print(f'Pasta já existe: {caminho_pasta}')
                
            inserirDadosExcel(codigo, caminho_pasta)
                
    elif filtro == 'FILTRO B':
        for codigo, nome in zip(regionaisCodigo1, regionaisNome1):
            nome_pasta = f'{codigo} - {nome}'
            caminho_pasta = os.path.join(caminho_mes, nome_pasta)
            if not os.path.exists(caminho_pasta):
                os.makedirs(caminho_pasta)
                print(f'Pasta criada: {caminho_pasta}')
            else:
                print(f'Pasta já existe: {caminho_pasta}')
                
            inserirDadosExcel(codigo, caminho_pasta)
    else:
        for codigo, nome in zip(regionaisCodigo2, regionaisNome2):
            nome_pasta = f'{codigo} - {nome}'
            caminho_pasta = os.path.join(caminho_mes, nome_pasta)
            if not os.path.exists(caminho_pasta):
                os.makedirs(caminho_pasta)
                print(f'Pasta criada: {caminho_pasta}')
            else:
                print(f'Pasta já existe: {caminho_pasta}')
                
            inserirDadosExcel(codigo, caminho_pasta)

def executarAutomacao():
    print('---- Iniciando Automação ----')
    
    tempo_total = 0
    tempo_total += criarPastasRegionais('PARAMETRO')
    tempo_total += criarPastasRegionais('PARAMETRO 1')
    tempo_total += criarPastasRegionais('PARAMETRO 2')
    
    print(f'Fim da Automação\nTempo de execução total de: {formatar_tempo(tempo_total)}')
    
executarAutomacao()
            
    